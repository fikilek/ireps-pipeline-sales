from __future__ import annotations

import copy
import hashlib
import json
import sys
import tempfile
import unittest
from datetime import UTC, datetime
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))
import sales_monthly_categories as categories
import sales_pipeline_sales_all_refresh as refresh
import sales_population_artifacts as population
from google.cloud.firestore_v1 import _helpers as real_helpers
from google.cloud.firestore_v1.types import document as real_document
from test_stage08_global_preflight_offline import google_module_stubs

NOW = datetime(2026, 9, 5, tzinfo=UTC)
CREATED = datetime(2026, 8, 8, tzinfo=UTC)
VALUE = {"leakageCategory": "CAT1", "riskTier": "LOW", "riskScore": 0}
ACTOR = {"uid": "approved-test-actor", "user": "Approved Test Actor"}


def fixture_package(directory):
    import openpyxl
    users = directory / "users.json"
    users.write_text(json.dumps({"users/" + ACTOR["uid"]: {"uid": ACTOR["uid"],
        "profile": {"displayName": ACTOR["user"]}, "__exists__": True},
        "users/" + categories.APPROVED_CREATOR["uid"]: {"uid": categories.APPROVED_CREATOR["uid"],
            "profile": {"displayName": categories.APPROVED_CREATOR["user"]}, "__exists__": True}}))
    ev = {"path": str(users), "sha256": categories.sha(users)}
    baseline = directory / "baseline.jsonl"
    baseline.write_bytes(b'{"meterNoNormalized":"M1"}\n')
    sources = []
    for month, sheet, column in (("2026-06", "Sheet1", "Leakage_Category"),
            ("2026-07", "Prepaid_30Month_Analysis", "July_2026_Category")):
        book = openpyxl.Workbook()
        book.active.title = sheet
        book.active.append(["MeterNumber", column, "Risk_Tier", "Risk_Score"])
        book.active.append(["M1", "CAT1", "LOW", 0])
        path = directory / (month + ".xlsx")
        book.save(path)
        book.close()
        sources.append({"path": str(path), "sha256": categories.sha(path), "sheet": sheet, "month": month})
    package = {"schemaVersion": 1, "month": "2026-07", "projectId": "ireps2",
        "source": sources[1], "historySources": [sources[0]],
        "historicalCategories": {"M1": {"2026-06": VALUE}},
        "creatorEvidence": ev, "actorEvidence": ev,
        "creatorScope": {"path": str(baseline), "sha256": categories.sha(baseline)},
        "creator": categories.APPROVED_CREATOR, "actor": ACTOR, "creatorEligibleIds": ["M1"], "categories": {"M1": VALUE}, "exceptions": []}
    return package


def payload(meter="M1"):
    return {"master": {"id": meter, "visibility": "VISIBLE"},
        "meterNoNormalized": meter, "provider": "contour", "lmPcode": "ZA5241",
        "leakageCategory": "frozen", "tbRefs": {"keep": True},
        "salesStatus": {"current": "UNCHANGED"},
        "monthlyCategories": {"2026-06": VALUE.copy()},
        "metadata": {"createdAt": CREATED, "createdByUid": "original",
            "createdByUser": "Original", "updatedAt": CREATED,
            "updatedByUid": "original", "updatedByUser": "Original"}}


def item(meter="M1", month="2026-07"):
    return {"masterId": meter, "expected": payload(meter), "categoryRefresh": {
        "month": month, "category": VALUE.copy(), "creator": ACTOR, "actor": ACTOR}}


class DB:
    def __init__(self, docs):
        self.docs = copy.deepcopy(docs)
        self.commits, self.reads = [], []

    def collection(self, name):
        return SimpleNamespace(document=lambda key: SimpleNamespace(id=key, path=f"{name}/{key}"))

    def get_all(self, refs):
        self.reads.append(len(refs))
        return [SimpleNamespace(id=r.id, reference=r, exists=r.id in self.docs,
            update_time=NOW, create_time=CREATED,
            to_dict=lambda key=r.id: copy.deepcopy(self.docs.get(key))) for r in refs]

    def batch(self):
        operations = []
        creates = []
        def commit():
            self.commits.append(copy.deepcopy(operations))
            for ref, value in creates:
                self.docs[ref.id] = copy.deepcopy(value)
            for ref, updates in operations:
                for path, value in updates.items():
                    target = self.docs[ref.id]
                    parts = path.split(".")
                    for part in parts[:-1]:
                        target = target.setdefault(part, {})
                    target[parts[-1]] = value
        return SimpleNamespace(update=lambda ref, updates, option: operations.append((ref, updates)),
            create=lambda ref, value: creates.append((ref, value)), commit=commit)

    def close(self):
        pass


class MonthlyCategoryTests(unittest.TestCase):
    def setUp(self):
        # Replace only the pinned production baseline digest with the exact
        # synthetic baseline's digest; identity and byte verification remain real.
        patcher = patch.object(categories, "APPROVED_CREATOR_SCOPE_SHA",
            hashlib.sha256(b'{"meterNoNormalized":"M1"}\n').hexdigest())
        patcher.start()
        self.addCleanup(patcher.stop)
    def test_exact_month_patch_and_preservation(self):
        doc = payload()
        updates = categories.changes(doc, item()["categoryRefresh"], CREATED, NOW)
        self.assertEqual(set(updates), {"monthlyCategories.2026-07", "metadata.updatedAt",
            "metadata.updatedByUid", "metadata.updatedByUser"})
        self.assertEqual(doc, payload())

    def test_identical_category_does_not_churn_metadata(self):
        self.assertEqual(categories.changes(payload(), item(month="2026-06")["categoryRefresh"], CREATED, NOW), {})

    def test_history_conflict_and_malformed_history_fail(self):
        context = item(month="2026-06")["categoryRefresh"]
        context["category"]["riskScore"] = 4
        with self.assertRaisesRegex(ValueError, "Historical category conflict"):
            categories.changes(payload(), context, CREATED, NOW)
        doc = payload()
        doc["monthlyCategories"]["bad-month"] = VALUE
        with self.assertRaisesRegex(ValueError, "malformed"):
            categories.changes(doc, item()["categoryRefresh"], CREATED, NOW)

    def test_invalid_risk_types(self):
        for score in (True, 1.0, "1", None, -1):
            with self.subTest(score=score), self.assertRaises(ValueError):
                categories.category(dict(VALUE, riskScore=score))

    def test_missing_creation_evidence_blocks(self):
        doc = payload()
        del doc["metadata"]
        with self.assertRaisesRegex(ValueError, "timestamp unavailable"):
            categories.changes(doc, item()["categoryRefresh"], None, NOW)
        updates = categories.changes(doc, item()["categoryRefresh"], CREATED, NOW)
        self.assertEqual(updates["metadata.createdAt"], CREATED)
        self.assertNotEqual(updates["metadata.createdAt"], NOW)
        context = item()["categoryRefresh"]
        context["creator"] = None
        with self.assertRaisesRegex(ValueError, "actor provenance unavailable"):
            categories.changes(doc, context, CREATED, NOW)

    def test_global_conflict_blocks_every_write_and_no_create(self):
        rows = [item("M1"), item("M2")]
        db = DB({"M1": payload("M1")})
        stats = refresh.RefreshStats(2)
        plan = refresh.classify_all(db=db, collection=db.collection(refresh.COLLECTION), rows=rows,
            stats=stats, preserved_before={})
        with self.assertRaises(RuntimeError):
            refresh.evaluate_global_gate(rows=rows, plan=plan, stats=stats)
        self.assertEqual(stats.conflicts, 1)
        self.assertEqual(stats.created, 0)
        self.assertEqual(db.commits, [])

    def test_sales_history_shrink_and_mutation_block(self):
        doc = payload()
        doc["monthlySalesC"] = {"2026-06": 1}
        self.assertIn("Historical Sales conflict", refresh._conflict(doc, payload()))
        changed = copy.deepcopy(doc)
        changed["monthlySalesC"]["2026-06"] = 2
        self.assertIn("Historical Sales conflict", refresh._conflict(doc, changed))

    def test_sales_month_additions_are_dotted_and_legacy_scalars_frozen(self):
        doc = payload()
        wanted = copy.deepcopy(doc)
        doc["monthlySalesC"] = {"2026-06": 1}
        wanted["monthlySalesC"] = {"2026-06": 1, "2026-07": 2}
        wanted["leakageCategory"] = "new scalar must not write"
        updates = refresh._updates(doc, wanted)
        self.assertEqual(updates["monthlySalesC.2026-07"], 2)
        self.assertNotIn("monthlySalesC", updates)
        self.assertNotIn("leakageCategory", updates)

    def test_actual_refresh_reporting_preflight_and_execution(self):
        for preflight in (True, False):
            with self.subTest(preflight=preflight), tempfile.TemporaryDirectory() as temp:
                directory = Path(temp)
                evidence = directory / "evidence.json"
                evidence.write_text("{}")
                package = fixture_package(directory)
                package_path = directory / "category.json"
                package_path.write_text(json.dumps(package))
                service = directory / "service.json"
                service.write_text('{"project_id":"ireps2"}')
                db = DB({"M1": payload()})
                modules, _ = google_module_stubs(db)
                modules["google.cloud.firestore_v1"]._helpers = real_helpers
                with patch.dict(sys.modules, modules), patch.object(refresh, "load_and_validate",
                        return_value=([{"masterId": "M1", "expected": payload()}], {})):
                    report_path = refresh.run_refresh(project_id="ireps2", confirm_project="ireps2",
                        service_account_path=service, input_path=evidence, manifest_path=evidence,
                        report_dir=directory, preflight_only=preflight,
                        category_package_path=package_path, category_package_sha256=categories.sha(package_path))
                report = json.loads(report_path.read_text())
                self.assertEqual(report["status"], "PASS")
                self.assertEqual(report["writeAttemptCount"], 0 if preflight else 1)
                self.assertEqual(report["globalPreflight"], {"complete": True, "gatePassed": True})
                self.assertEqual(db.docs["M1"]["metadata"]["createdAt"], CREATED)
                self.assertEqual(db.docs["M1"]["tbRefs"], payload()["tbRefs"])
                self.assertEqual(db.docs["M1"]["monthlyCategories"]["2026-06"], VALUE)
                if not preflight:
                    self.assertEqual(report["verification"]["status"], "PASS")
                    self.assertEqual(db.docs["M1"]["monthlyCategories"]["2026-07"], VALUE)

    def test_snapshot_bytes_immutable_and_members_distinct(self):
        snapshot = population.build_snapshot(lm_pcode="ZA5241", provider="contour", month="2026-06",
            source_sha256="a"*64, members=["M1"], evidence_sha256="b"*64)
        digest = hashlib.sha256(population.encode(snapshot)).hexdigest()
        with tempfile.TemporaryDirectory() as directory:
            path, actual = population.write_snapshot(directory, snapshot)
            self.assertEqual(actual, digest)
            self.assertEqual(path.read_bytes(), population.encode(snapshot))
            self.assertEqual(population.write_snapshot(directory, snapshot), (path, digest))
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / f"{digest}.json"
            path.write_bytes(b"existing different bytes")
            with self.assertRaisesRegex(ValueError, "already exists with different bytes"):
                population.write_snapshot(directory, snapshot)
            self.assertEqual(path.read_bytes(), b"existing different bytes")
        with self.assertRaises(ValueError):
            population.build_snapshot(lm_pcode="ZA5241", provider="contour", month="2026-06",
                source_sha256="a"*64, members=["M1", "M1"], evidence_sha256="b"*64)

    def test_forged_actor_creator_scope_or_category_cannot_enter_preflight(self):
        for mutation, error in (("actor", "does not match"), ("scope", "eligibility differs"),
                ("category", "differs from authoritative"), ("history", "Historical category package")):
            with self.subTest(mutation=mutation), tempfile.TemporaryDirectory() as temp:
                directory = Path(temp)
                package = fixture_package(directory)
                if mutation == "actor":
                    package["actor"] = {"uid": "forged", "user": "Forged"}
                elif mutation == "scope":
                    package["creatorEligibleIds"] = ["OTHER"]
                elif mutation == "category":
                    package["categories"] = {"M1": dict(VALUE, riskScore=99)}
                else:
                    package["historicalCategories"] = {"M1": {}}
                path = directory / "package.json"
                path.write_text(json.dumps(package))
                with self.assertRaisesRegex(ValueError, error):
                    categories.load_package(path, categories.sha(path), [item()], "ireps2")

    def test_prior_governed_month_missing_blocks_next_month(self):
        row = item()
        row["categoryRefresh"]["requiredHistory"] = {"2026-06": VALUE}
        doc = payload()
        del doc["monthlyCategories"]
        db = DB({"M1": doc})
        snapshot = db.get_all([db.collection(refresh.COLLECTION).document("M1")])[0]
        decision = refresh._classify_snapshot(row, snapshot)
        self.assertEqual(decision["classification"], "CONFLICT")
        self.assertIn("Required historical category", decision["reason"])

    def test_future_refresh_cannot_skip_intermediate_governed_month(self):
        import openpyxl
        with tempfile.TemporaryDirectory() as temp:
            directory = Path(temp)
            package = fixture_package(directory)
            source = directory / "september.xlsx"
            book = openpyxl.Workbook()
            book.active.title = "Prepaid_30Month_Analysis"
            book.active.append(["MeterNumber", "September_2026_Category", "Risk_Tier", "Risk_Score"])
            book.active.append(["M1", "CAT1", "LOW", 0])
            book.save(source)
            book.close()
            package["month"] = "2026-09"
            package["source"] = {"path": str(source), "sha256": categories.sha(source), "sheet": "Prepaid_30Month_Analysis"}
            path = directory / "package.json"
            path.write_text(json.dumps(package))
            with self.assertRaisesRegex(ValueError, "Complete June-to-target"):
                categories.load_package(path, categories.sha(path), [item()], "ireps2")
        self.assertEqual(categories.required_history_months("2026-09"), ["2026-06", "2026-07", "2026-08"])
        with self.assertRaises(ValueError):
            categories.required_history_months("2026-05")

    def test_missing_category_is_genuinely_classified_without_fabricated_metadata(self):
        with tempfile.TemporaryDirectory() as temp:
            directory = Path(temp)
            package = fixture_package(directory)
            package["historicalCategories"]["M2"] = {}
            package["exceptions"] = [{"meterId": "M2", "reason": "No authoritative exact-month category source row"}]
            path = directory / "package.json"
            path.write_text(json.dumps(package))
            rows, evidence = categories.load_package(path, categories.sha(path), [item("M1"), item("M2")], "ireps2")
            doc = payload("M2")
            del doc["metadata"]
            db = DB({"M1": payload(), "M2": doc})
            stats = refresh.RefreshStats(2)
            plan = refresh.classify_all(db=db, collection=db.collection(refresh.COLLECTION), rows=rows,
                stats=stats, preserved_before={})
            refresh.evaluate_global_gate(rows=rows, plan=plan, stats=stats)
            self.assertEqual(stats.inspected, 2)
            self.assertEqual(stats.updated, 1)
            self.assertEqual(stats.unchanged, 1)
            self.assertEqual(plan[0]["decisions"][1]["updates"], {})
            self.assertEqual(evidence["categoryRecords"], 1)

    def test_missing_category_backfills_only_approved_creation_metadata(self):
        doc = payload()
        del doc["metadata"]
        context = item(month="2026-06")["categoryRefresh"]
        context["category"] = None
        before = copy.deepcopy(doc)
        updates = categories.changes(doc, context, CREATED, NOW)
        self.assertEqual(set(updates), {f"metadata.{key}" for key in categories.META})
        self.assertEqual(updates["metadata.createdAt"], CREATED)
        self.assertEqual(updates["metadata.createdByUid"], ACTOR["uid"])
        self.assertEqual(doc, before)
        with self.assertRaisesRegex(ValueError, "Original Firestore creation timestamp"):
            categories.changes(doc, context, None, NOW)
        context["creator"] = None
        self.assertEqual(categories.changes(doc, context, CREATED, NOW), {})

    def test_user_attribution_requires_matching_identity_scope_and_hashes(self):
        with tempfile.TemporaryDirectory() as temp:
            directory = Path(temp)
            scope = directory / "scope.csv"
            scope.write_text("masterId\nM1\nM2\n")
            confirmation = directory / "confirmation.txt"
            confirmation.write_text("Synthetic user attribution fixture")
            policy = {"schemaVersion": 1, "projectId": "ireps2",
                "authorityType": "USER_CONFIRMED_PIPELINE_ATTRIBUTION", "identity": ACTOR,
                "existingDocumentIds": ["M1", "M2"],
                "sourceScope": {"path": str(scope), "sha256": categories.sha(scope)},
                "confirmation": {"path": str(confirmation), "sha256": categories.sha(confirmation)}}
            path = directory / "policy.json"
            def contract():
                path.write_text(json.dumps(policy))
                return {"creatorEligibleIds": ["M1"], "creator": ACTOR, "actor": ACTOR,
                    "pipelineAttributionEvidence": {"path": str(path), "sha256": categories.sha(path)}}
            self.assertEqual(categories.creator_eligible_ids(contract(), "ireps2"), {"M1", "M2"})
            with self.assertRaisesRegex(ValueError, "identity/project"):
                categories.creator_eligible_ids(contract(), "other-project")
            policy["existingDocumentIds"].append("M3")
            with self.assertRaisesRegex(ValueError, "scope mismatch"):
                categories.creator_eligible_ids(contract(), "ireps2")
            policy["existingDocumentIds"].pop()
            pinned = contract()
            confirmation.write_text("Changed evidence")
            with self.assertRaisesRegex(ValueError, "hash mismatch"):
                categories.creator_eligible_ids(pinned, "ireps2")

    def test_partial_metadata_preserves_creation_and_flags_contradiction(self):
        context = {"creator": ACTOR, "actor": ACTOR, "pipelineAttributionConfirmed": True}
        doc = {"metadata": {"createdAt": CREATED, "createdByUid": ACTOR["uid"]}}
        updates = categories.metadata_patch(doc, context, None, NOW)
        self.assertNotIn("metadata.createdAt", updates)
        self.assertNotIn("metadata.createdByUid", updates)
        self.assertEqual(updates["metadata.createdByUser"], ACTOR["user"])
        complete = {"metadata": dict(doc["metadata"], **{k.split('.')[1]: v for k, v in updates.items()})}
        self.assertEqual(categories.metadata_patch(complete, context, None, NOW), {})
        complete["metadata"]["createdByUid"] = "different-existing-creator"
        with self.assertRaisesRegex(ValueError, "contradicts confirmed"):
            categories.metadata_patch(complete, context, CREATED, NOW)

    def test_commercial_ingestion_preserves_complete_history(self):
        original = [{"meterNoNormalized": "M1", "monthlyCategories": {"2026-06": VALUE}, "keep": 1}]
        added = categories.append_to_commercial(original, "2026-07", {"M1": VALUE})
        self.assertEqual(set(added[0]["monthlyCategories"]), {"2026-06", "2026-07"})
        self.assertEqual(original[0]["monthlyCategories"], {"2026-06": VALUE})
        with self.assertRaises(ValueError):
            categories.append_to_commercial(original, "2026-06", {"M1": dict(VALUE, riskScore=2)})
        with self.assertRaises(ValueError):
            categories.append_to_commercial(original, "2026-07", {"M2": VALUE})

    def test_stage05_06_08_roundtrip_complete_category_map(self):
        import pandas as pd
        import sales_pipeline_monthly_source_support as support
        from test_stage06_stage08_offline import synthetic_commercial_row, write_jsonl
        with tempfile.TemporaryDirectory() as temp:
            directory = Path(temp)
            commercial = directory / "commercial.jsonl"
            row = synthetic_commercial_row()
            row["monthlyCategories"] = {"2026-06": VALUE}
            source_sha = write_jsonl(commercial, [row])
            monthly = directory / "monthly"
            monthly.mkdir()
            pd.DataFrame([{"lmPcode": "ZA7423", "meterNo": "ABC123", "ym": "2026-06",
                "provider": "contour", "amountTotalC": 100, "unitsTotal": 5,
                "sourceOrigin": "monthly_source"}]).to_csv(
                    monthly / "monthly__FULL__2026-06__from_monthly_source.csv", index=False)
            master = directory / "master.csv"
            master_manifest = directory / "master.manifest.json"
            support.build_stage05_monthly_source(lm_pcode="ZA7423", provider="contour",
                meter_type="electricity", from_month="2026-06", to_month="2026-06",
                commercial_source=commercial, expected_commercial_sha256=source_sha,
                monthly_dir=monthly, output_path=master, manifest_path=master_manifest, source_run_id="TEST")
            sales = directory / "sales.csv"
            sales_manifest = directory / "sales.manifest.json"
            support.build_stage06_monthly_source(lm_pcode="ZA7423", provider="contour",
                from_month="2026-06", to_month="2026-06", master_path=master,
                master_manifest_path=master_manifest, commercial_source=commercial,
                expected_commercial_sha256=source_sha, monthly_dir=monthly,
                output_path=sales, manifest_path=sales_manifest)
            loaded, _ = refresh.load_and_validate(sales, sales_manifest)
            self.assertEqual(loaded[0]["expected"]["monthlyCategories"], {"2026-06": VALUE})

    def test_actual_stage03a_cli_appends_exact_month_category(self):
        import subprocess
        from test_stage03a_monthly_source_refresh_offline import baseline_record, make_xlsx, wb_row
        with tempfile.TemporaryDirectory() as temp:
            directory = Path(temp)
            contract = fixture_package(directory)
            baseline = directory / "commercial.jsonl"
            row = baseline_record("M1")
            row["monthlyCategories"] = {"2026-06": VALUE}
            baseline.write_text(json.dumps(row) + "\n")
            supplier = directory / "supplier.xlsx"
            make_xlsx(supplier, [wb_row("M1", sales={"2026-06": "100", "2026-07": "120"},
                units={"2026-06": "20", "2026-07": "25"})], ["2026-06", "2026-07"])
            output = directory / "output.jsonl"
            report = directory / "report.json"
            result = subprocess.run([sys.executable, "-B", str(Path(__file__).resolve().parents[1] / "scripts/03a_refresh_monthly_source_commercial.py"),
                "--baseline", str(baseline), "--expected-baseline-sha256", categories.sha(baseline),
                "--bootstrap-previous-from-baseline", "--workbook", str(supplier), "--expected-workbook-sha256", categories.sha(supplier),
                "--sheet", "Purchases", "--lm-pcode", "ZA5241", "--provider", "contour",
                "--from-month", "2026-07", "--to-month", "2026-07", "--source-run-id", "TEST",
                "--report", str(report), "--write", "--output", str(output), "--snapshot-output", str(directory / "snapshot.json"),
                "--category-source", contract["source"]["path"], "--expected-category-source-sha256", contract["source"]["sha256"],
                "--category-sheet", contract["source"]["sheet"]], capture_output=True, text=True, timeout=30)
            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            loaded = json.loads(output.read_text().strip())
            self.assertEqual(loaded["monthlyCategories"], {"2026-06": VALUE, "2026-07": VALUE})
            self.assertEqual(loaded["monthlySalesC"]["2026-06"], 10000)
            self.assertEqual(json.loads(report.read_text())["categorySource"]["populationAuthority"], False)

    def test_standard_commercial_creation_and_noop_metadata(self):
        for exists in (False, True):
            with self.subTest(exists=exists), tempfile.TemporaryDirectory() as temp:
                directory = Path(temp)
                contract = fixture_package(directory)
                path = directory / "metadata.json"
                path.write_text(json.dumps(contract))
                service = directory / "service.json"
                service.write_text('{"project_id":"ireps2"}')
                db = DB({"M1": payload()} if exists else {})
                modules, _ = google_module_stubs(db)
                modules["google.cloud.firestore_v1"]._helpers = real_helpers
                with patch.dict(sys.modules, modules), patch.object(refresh, "load_and_validate",
                        return_value=([{"masterId": "M1", "expected": payload()}], {})):
                    report = refresh.run_refresh(project_id="ireps2", confirm_project="ireps2",
                        service_account_path=service, input_path=path, manifest_path=path,
                        report_dir=directory, preflight_only=False,
                        metadata_contract_path=path, metadata_contract_sha256=categories.sha(path))
                report = json.loads(report.read_text())
                self.assertEqual(report["writeSuccessCount"], 0 if exists else 1)
                self.assertEqual(report["recoveryEvidence"]["complete"], True)
                if exists:
                    self.assertEqual(db.docs["M1"]["metadata"]["updatedAt"], CREATED)
                else:
                    self.assertEqual(db.docs["M1"]["metadata"]["createdByUid"], ACTOR["uid"])
                    self.assertNotIn("leakageCategory", db.docs["M1"])

    def test_legacy_cli_write_modes_and_unattributed_refresh_are_blocked(self):
        from test_stage06_stage08_offline import stage08
        for mode in ("create-only", "initial-load", "resume"):
            with self.subTest(mode=mode), patch.object(stage08, "parse_args",
                    return_value=SimpleNamespace(mode=mode, preflight_only=False)):
                with self.assertRaisesRegex(ValueError, "Final-state Sales writes require"):
                    stage08.main()
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "sa.json"
            path.write_text('{"project_id":"ireps2"}')
            with patch.object(refresh, "load_and_validate", return_value=([item()], {})):
                with self.assertRaisesRegex(ValueError, "requires an approved metadata contract"):
                    refresh.run_refresh(project_id="ireps2", confirm_project="ireps2",
                        service_account_path=path, input_path=path, manifest_path=path,
                        report_dir=Path(temp), preflight_only=False)

    def test_stage08_cli_forwards_category_preflight_flag(self):
        from test_stage06_stage08_offline import stage08
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp)
            argv = ["08_upload_sales_all_meters.py", "--project-id", "ireps2", "--confirm-project", "ireps2",
                "--service-account", str(path / "sa.json"), "--input", str(path / "sales.csv"),
                "--manifest", str(path / "manifest.json"), "--mode", "refresh", "--preflight-only",
                "--category-package", str(path / "package.json"), "--category-package-sha256", "a"*64]
            with patch.object(sys, "argv", argv), patch.object(refresh, "run_refresh", return_value=path / "report.json") as execute:
                stage08.main()
            self.assertIs(execute.call_args.kwargs["preflight_only"], True)
            self.assertEqual(execute.call_args.kwargs["category_package_path"], path / "package.json")
            self.assertEqual(execute.call_args.kwargs["category_package_sha256"], "a"*64)


class CorrectedJulyAdmissionTests(unittest.TestCase):
    def workbook(self, directory, identities):
        import openpyxl
        book = openpyxl.Workbook()
        book.active.title = "Prepaid_30Month_Analysis"
        book.active.append(["MeterNumber", "PreviousMeterNumber", "CorrectedMeterNumber",
            "July_2026_Category", "Risk_Tier", "Risk_Score"])
        for value in identities:
            book.active.append(["00000000001", "00000000002", value, "CAT1", "LOW", 0])
        path = directory / "corrected.xlsx"
        book.save(path)
        book.close()
        return path

    def test_configured_corrected_text_and_13_char_identity(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(Path(temp), ["0261249145812"])
            evidence = {"0261249145812": None}
            values, exceptions, aliases = categories.ingest_workbook(path, categories.sha(path),
                "Prepaid_30Month_Analysis", "2026-07", set(evidence),
                identity_field="CorrectedMeterNumber", source_rows=evidence)
            self.assertEqual(set(values), set(evidence))
            self.assertEqual((exceptions, aliases), ([], []))
            self.assertEqual(evidence["0261249145812"]["PreviousMeterNumber"], "00000000002")

    def test_legacy_default_ignores_corrected_column(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(Path(temp), ["0261249145812"])
            values, _, _ = categories.ingest_workbook(path, categories.sha(path),
                "Prepaid_30Month_Analysis", "2026-07", {"00000000001"})
            self.assertEqual(set(values), {"00000000001"})

    def test_blank_numeric_malformed_corrected_fail_closed(self):
        for invalid in (None, "", 4297839708, " 04297839708", "04-297839708", "00000000000"):
            with self.subTest(invalid=invalid), tempfile.TemporaryDirectory() as temp:
                path = self.workbook(Path(temp), [invalid])
                with self.assertRaisesRegex(ValueError, "Blank/invalid"):
                    categories.ingest_workbook(path, categories.sha(path), "Prepaid_30Month_Analysis",
                        "2026-07", {"04297839708"}, identity_field="CorrectedMeterNumber")

    def test_duplicate_corrected_fails_closed(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(Path(temp), ["04297839708", "04297839708"])
            with self.assertRaisesRegex(ValueError, "Duplicate classification"):
                categories.ingest_workbook(path, categories.sha(path), "Prepaid_30Month_Analysis",
                    "2026-07", {"04297839708"}, identity_field="CorrectedMeterNumber")

    def test_corrected_never_pads_or_uses_nearest_id(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(Path(temp), ["4297839708"])
            with self.assertRaisesRegex(ValueError, "outside execution population"):
                categories.ingest_workbook(path, categories.sha(path), "Prepaid_30Month_Analysis",
                    "2026-07", {"04297839708"}, identity_field="CorrectedMeterNumber")

    def test_corrected_configuration_cannot_change_june(self):
        with tempfile.TemporaryDirectory() as temp:
            path = self.workbook(Path(temp), ["04297839708"])
            with self.assertRaisesRegex(ValueError, "governed July"):
                categories.ingest_workbook(path, categories.sha(path), "Sheet1",
                    "2026-06", {"04297839708"}, identity_field="CorrectedMeterNumber")

    def test_broad_creator_evidence_does_not_expand_execution(self):
        with tempfile.TemporaryDirectory() as temp:
            directory = Path(temp)
            package = fixture_package(directory)
            baseline = Path(package["creatorScope"]["path"])
            baseline.write_text('{"meterNoNormalized":"M1"}\n{"meterNoNormalized":"EXITED"}\n')
            package["creatorScope"]["sha256"] = categories.sha(baseline)
            package["creatorEligibleIds"] = ["M1", "EXITED"]
            package.update(executionIds=["M1"], lmPcode="ZA5241", provider="contour")
            snap = directory / "population.json"
            snap.write_text(json.dumps({"schemaVersion": 1, "month": "2026-07", "members": ["M1"],
                "lmPcode": "ZA5241", "provider": "contour", "sourceSha256": package["source"]["sha256"],
                "completeness": {"complete": True}}))
            package["populationSnapshot"] = {"path": str(snap), "sha256": categories.sha(snap)}
            path = directory / "package.json"
            path.write_text(json.dumps(package))
            with patch.object(categories, "APPROVED_CREATOR_SCOPE_SHA", categories.sha(baseline)):
                rows, evidence = categories.load_package(path, categories.sha(path),
                    [item(), item("EXITED"), item("OTHER")], "ireps2")
            self.assertEqual([r["masterId"] for r in rows], ["M1"])
            self.assertEqual(evidence["exactExecutionIds"], ["M1"])
            decisions = [refresh._classify_snapshot(r, SimpleNamespace(exists=True,
                create_time=CREATED, to_dict=payload)) for r in rows]
            self.assertEqual([d["masterId"] for d in decisions], ["M1"])
            self.assertNotIn("EXITED", {d["masterId"] for d in decisions})

    def test_creation_uses_normal_metadata_and_create_batch(self):
        row = {"masterId": "NEW", "createOnly": True, "expected": payload("NEW"),
            "metadataRefresh": {"actor": ACTOR, "creator": ACTOR}}
        row["expected"].pop("metadata")
        row["expected"]["monthlyCategories"] = {"2026-07": VALUE}
        decision = refresh._classify_snapshot(row, None)
        self.assertEqual(decision["classification"], "CREATED")
        self.assertEqual(set(row["expected"]["metadata"]), categories.META)
        self.assertIsInstance(row["expected"]["metadata"]["createdAt"], datetime)
        self.assertEqual(row["expected"]["metadata"]["createdByUid"], ACTOR["uid"])
        self.assertNotIn("leakageCategory", row["expected"])
        self.assertEqual(set(row["expected"]["monthlyCategories"]), {"2026-07"})
        from unittest.mock import Mock
        db = Mock()
        collection = Mock()
        refresh._build_write_batch(db=db, collection=collection,
            operations=[dict(decision, expected=row["expected"])], last_update_option_cls=Mock())
        db.batch.return_value.create.assert_called_once()
        db.batch.return_value.update.assert_not_called()
        db.batch.return_value.commit.assert_not_called()

    def test_creation_admission_preserves_source_blanks_and_populated_units(self):
        import openpyxl
        with tempfile.TemporaryDirectory() as temp:
            directory = Path(temp)
            package = fixture_package(directory)
            path = Path(package["source"]["path"])
            book = openpyxl.load_workbook(path)
            for col, name in enumerate(["2026-06", "2026-07", "2026-06.1", "2026-07.1"], 5):
                book.active.cell(1, col, name)
            book.active.append(["NEW", "CAT1", "LOW", 0, None, 0, None, 100])
            book.save(path)
            book.close()
            package["source"]["sha256"] = categories.sha(path)
            package.update(executionIds=["M1", "NEW"], lmPcode="ZA5241", provider="contour")
            package["categories"]["NEW"] = VALUE
            package["historicalCategories"]["NEW"] = {}
            snapshot = directory / "snapshot.json"
            snapshot.write_text(json.dumps({"schemaVersion": 1, "month": "2026-07", "members": ["M1", "NEW"],
                "lmPcode": "ZA5241", "provider": "contour", "sourceSha256": categories.sha(path),
                "completeness": {"complete": True}}))
            package["populationSnapshot"] = {"path": str(snapshot), "sha256": categories.sha(snapshot)}
            placeholder = directory / "stage06.json"
            placeholder.write_text('{}')
            ref = {"path": str(placeholder), "sha256": categories.sha(placeholder)}
            package["creationStage06"] = {"input": ref, "manifest": ref, "ids": ["NEW"]}
            expected = {"master": {"id": "NEW", "visibility": "INVISIBLE"},
                "meterNo": "NEW", "meterNoNormalized": "NEW", "lmPcode": "ZA5241", "provider": "contour",
                "monthlyCategories": {"2026-07": VALUE}, "salesPeriodTo": "2026-07", "sourceEndRow": 3,
                "monthlySalesC": {"2026-06": 0, "2026-07": 0}, "monthlyUnits": {"2026-06": 0, "2026-07": 100},
                "totalSalesC": 0, "totalUnits": 100}
            pkg_path = directory / "package.json"
            pkg_path.write_text(json.dumps(package))
            with patch.object(categories, "APPROVED_CREATOR_SCOPE_SHA", package["creatorScope"]["sha256"]), \
                    patch.object(refresh, "load_and_validate", return_value=([{"masterId": "NEW", "expected": expected}], {})):
                rows, _ = categories.load_package(pkg_path, categories.sha(pkg_path), [item()], "ireps2")
            new = next(r for r in rows if r["masterId"] == "NEW")
            self.assertEqual(new["expected"]["monthlySalesC"], {"2026-07": 0})
            self.assertEqual(new["expected"]["monthlyUnits"], {"2026-07": 100})
            self.assertNotIn("categoryRefresh", new)
            self.assertTrue(new["createOnly"])
            # A blank source is not a zero. No-history members retain empty
            # required maps and zero aggregate totals without an invented month.
            book = openpyxl.load_workbook(path)
            book.active.cell(3, 6).value = None
            book.active.cell(3, 8).value = None
            book.save(path)
            book.close()
            package["source"]["sha256"] = categories.sha(path)
            snap = json.loads(snapshot.read_text())
            snap["sourceSha256"] = categories.sha(path)
            snapshot.write_text(json.dumps(snap))
            package["populationSnapshot"]["sha256"] = categories.sha(snapshot)
            pkg_path.write_text(json.dumps(package))
            expected["monthlyUnits"]["2026-07"] = 0
            expected["totalUnits"] = 0
            with patch.object(categories, "APPROVED_CREATOR_SCOPE_SHA", package["creatorScope"]["sha256"]), \
                    patch.object(refresh, "load_and_validate", return_value=([{"masterId": "NEW", "expected": expected}], {})):
                rows, _ = categories.load_package(pkg_path, categories.sha(pkg_path), [item()], "ireps2")
            new = next(r for r in rows if r["masterId"] == "NEW")
            for field in ("monthlySalesC", "monthlyTotalsC", "monthlyUnits"):
                self.assertEqual(new["expected"][field], {})
            self.assertEqual(new["expected"]["totalSalesC"], 0)
            self.assertEqual(new["expected"]["totalUnits"], 0)

    def test_planned_creation_now_existing_is_conflict(self):
        row = {"masterId": "NEW", "createOnly": True, "expected": payload("NEW")}
        decision = refresh._classify_snapshot(row, SimpleNamespace(exists=True))
        self.assertEqual(decision["classification"], "CONFLICT")
        self.assertEqual(decision["updates"], {})


if __name__ == "__main__":
    unittest.main()
