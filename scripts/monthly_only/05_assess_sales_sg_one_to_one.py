#!/usr/bin/env python3
"""
Read-only one-to-one assessment for the Endumeni monthly-only sales PSD.

Links:
END MeterNumber -> END AccountNumber
-> CSM valuation roll OWNER_ACCOUNT_NO -> GIS_KEY
-> authoritative ERF pipeline SG/parcel key

This script performs no Firestore operations and does not overwrite the PSD.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PROGRESS_EVERY = 2_000
AUTO_DETECT_SAMPLE_LINES = 2_000
K241_PREFIX = "K241"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--end-xlsx", required=True, type=Path)
    parser.add_argument("--bridge-xlsx", required=True, type=Path)
    parser.add_argument("--erf-jsonl", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--expected-end-count", type=int, default=10_216)
    parser.add_argument(
        "--erf-key-path",
        default="",
        help=(
            "Optional dotted JSON path containing the ERF SG/parcel key. "
            "When omitted, the script auto-detects the best path by overlap "
            "with valuation-roll GIS_KEY values."
        ),
    )
    return parser.parse_args()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def require_file(path: Path, label: str) -> None:
    if not path.is_file():
        raise SystemExit(f"{label} not found: {path}")


def normalize_account(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    if text.startswith("'"):
        text = text[1:].strip()

    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]

    return text


def normalize_meter(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def normalize_sg_key(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    if text.startswith("'"):
        text = text[1:].strip()

    text = re.sub(r"[^A-Za-z0-9]", "", text).upper()

    if text.startswith(K241_PREFIX):
        text = text[len(K241_PREFIX) :]

    return text


def get_header_index(headers: list[Any], expected: str) -> int:
    normalized = {
        str(value).strip().upper(): index
        for index, value in enumerate(headers)
        if value is not None
    }

    key = expected.strip().upper()
    if key not in normalized:
        raise SystemExit(
            f"Required column {expected!r} was not found. "
            f"Available columns: {sorted(normalized)}"
        )

    return normalized[key]


def flatten_scalars(value: Any, prefix: str = "") -> Iterable[tuple[str, Any]]:
    if isinstance(value, dict):
        for key, child in value.items():
            child_path = f"{prefix}.{key}" if prefix else str(key)
            yield from flatten_scalars(child, child_path)
        return

    if isinstance(value, list):
        return

    yield prefix, value


def get_path_value(value: Any, dotted_path: str) -> Any:
    current = value
    for segment in dotted_path.split("."):
        if not isinstance(current, dict) or segment not in current:
            return None
        current = current[segment]
    return current


def read_end_rows(path: Path) -> tuple[list[dict[str, Any]], Counter]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active

    headers = list(
        next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    )
    meter_index = get_header_index(headers, "MeterNumber")
    account_index = get_header_index(headers, "AccountNumber")

    rows: list[dict[str, Any]] = []
    meters_seen: Counter = Counter()

    for source_row, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        meter_number = normalize_meter(row[meter_index])
        account_number = normalize_account(row[account_index])

        if not meter_number:
            raise SystemExit(f"Blank MeterNumber at END source row {source_row}")

        meters_seen[meter_number] += 1
        rows.append(
            {
                "sourceRow": source_row,
                "meterNumber": meter_number,
                "accountNumber": account_number,
            }
        )

        if len(rows) % PROGRESS_EVERY == 0:
            print(f"[PROGRESS] Read {len(rows):,} END meter rows")

    return rows, meters_seen


def read_bridge(
    path: Path,
) -> tuple[set[str], dict[str, set[str]], dict[str, list[int]], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active

    headers = list(
        next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    )
    account_index = get_header_index(headers, "OWNER_ACCOUNT_NO")
    gis_key_index = get_header_index(headers, "GIS_KEY")

    all_accounts: set[str] = set()
    account_to_sg: dict[str, set[str]] = defaultdict(set)
    relationship_rows: dict[str, list[int]] = defaultdict(list)
    data_rows = 0

    for source_row, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        data_rows += 1
        account_number = normalize_account(row[account_index])
        sg_key = normalize_sg_key(row[gis_key_index])

        if account_number:
            all_accounts.add(account_number)

        if account_number and sg_key:
            account_to_sg[account_number].add(sg_key)
            relationship_rows[f"{account_number}|{sg_key}"].append(source_row)

        if data_rows % PROGRESS_EVERY == 0:
            print(f"[PROGRESS] Read {data_rows:,} valuation-roll rows")

    return all_accounts, account_to_sg, relationship_rows, data_rows


def detect_erf_key_path(
    path: Path,
    bridge_sg_keys: set[str],
) -> tuple[str, list[dict[str, Any]]]:
    path_stats: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "nonblank": 0,
            "normalized": set(),
            "overlap": set(),
            "samples": [],
        }
    )

    sampled = 0
    with path.open("r", encoding="utf-8") as handle:
        for line_number, raw_line in enumerate(handle, start=1):
            if sampled >= AUTO_DETECT_SAMPLE_LINES:
                break

            raw_line = raw_line.strip()
            if not raw_line:
                continue

            try:
                record = json.loads(raw_line)
            except json.JSONDecodeError as error:
                raise SystemExit(
                    f"Invalid JSON at ERF JSONL line {line_number}: {error}"
                ) from error

            sampled += 1

            for scalar_path, scalar_value in flatten_scalars(record):
                if not isinstance(scalar_value, (str, int)):
                    continue

                normalized = normalize_sg_key(scalar_value)
                if not normalized:
                    continue

                stats = path_stats[scalar_path]
                stats["nonblank"] += 1
                stats["normalized"].add(normalized)

                if normalized in bridge_sg_keys:
                    stats["overlap"].add(normalized)
                elif normalized.endswith("0") and normalized[:-1] in bridge_sg_keys:
                    stats["overlap"].add(normalized[:-1])

                if len(stats["samples"]) < 3:
                    stats["samples"].append(str(scalar_value))

    ranked = []
    for scalar_path, stats in path_stats.items():
        ranked.append(
            {
                "path": scalar_path,
                "overlapCount": len(stats["overlap"]),
                "distinctNormalized": len(stats["normalized"]),
                "nonblank": stats["nonblank"],
                "samples": stats["samples"],
            }
        )

    ranked.sort(
        key=lambda item: (
            item["overlapCount"],
            item["nonblank"],
            item["distinctNormalized"],
        ),
        reverse=True,
    )

    if not ranked or ranked[0]["overlapCount"] == 0:
        raise SystemExit(
            "Could not auto-detect the ERF SG/parcel-key field. "
            "Rerun with --erf-key-path after inspecting the JSONL shape."
        )

    return ranked[0]["path"], ranked[:10]


def read_erf_keys(
    path: Path,
    key_path: str,
) -> tuple[dict[str, int], dict[str, list[str]], int]:
    key_counts: dict[str, int] = Counter()
    key_samples: dict[str, list[str]] = defaultdict(list)
    record_count = 0

    with path.open("r", encoding="utf-8") as handle:
        for line_number, raw_line in enumerate(handle, start=1):
            raw_line = raw_line.strip()
            if not raw_line:
                continue

            try:
                record = json.loads(raw_line)
            except json.JSONDecodeError as error:
                raise SystemExit(
                    f"Invalid JSON at ERF JSONL line {line_number}: {error}"
                ) from error

            record_count += 1
            raw_key = get_path_value(record, key_path)
            normalized_key = normalize_sg_key(raw_key)

            if normalized_key:
                key_counts[normalized_key] += 1

                if len(key_samples[normalized_key]) < 3:
                    doc_id = (
                        record.get("docId")
                        or record.get("id")
                        or (
                            record.get("data", {}).get("id")
                            if isinstance(record.get("data"), dict)
                            else None
                        )
                        or f"line:{line_number}"
                    )
                    key_samples[normalized_key].append(str(doc_id))

            if record_count % PROGRESS_EVERY == 0:
                print(f"[PROGRESS] Read {record_count:,} ERF pipeline records")

    return key_counts, key_samples, record_count


def resolve_bridge_sg_to_erf(
    bridge_sg_key: str,
    erf_key_counts: dict[str, int],
) -> tuple[str, int, str]:
    """
    Resolve one valuation-roll GIS_KEY against normalized ERF pipeline keys.

    Proven Endumeni comparison formats:
    - Valuation roll: N0GT01170000072400000
    - ERF pipeline after removing K241: N0GT011700000724000000

    The valuation key is therefore tested unchanged and with exactly one
    comparison-only trailing zero. Source values are never changed.
    """
    if not bridge_sg_key:
        return "", 0, "NONE"

    matching_candidates: list[tuple[str, str, int]] = []

    exact_count = int(erf_key_counts.get(bridge_sg_key, 0))
    if exact_count:
        matching_candidates.append(
            ("EXACT_FORMAT", bridge_sg_key, exact_count)
        )

    trailing_zero_key = f"{bridge_sg_key}0"
    trailing_zero_count = int(erf_key_counts.get(trailing_zero_key, 0))
    if trailing_zero_count:
        matching_candidates.append(
            (
                "APPEND_ONE_TRAILING_ZERO",
                trailing_zero_key,
                trailing_zero_count,
            )
        )

    if not matching_candidates:
        return "", 0, "NO_MATCH"

    total_matches = sum(item[2] for item in matching_candidates)

    if len(matching_candidates) == 1:
        rule, matched_key, _ = matching_candidates[0]
        return matched_key, total_matches, rule

    return (
        "|".join(item[1] for item in matching_candidates),
        total_matches,
        "MULTIPLE_COMPARISON_FORMATS",
    )


def main() -> int:
    args = parse_args()

    require_file(args.end_xlsx, "END workbook")
    require_file(args.bridge_xlsx, "Valuation-roll bridge workbook")
    require_file(args.erf_jsonl, "Authoritative ERF pipeline JSONL")

    args.output_dir.mkdir(parents=True, exist_ok=True)

    print("")
    print("============================================================")
    print("ENDUMENI SALES ONE-TO-ONE READ-ONLY ASSESSMENT")
    print("============================================================")
    print(f"END workbook:       {args.end_xlsx.resolve()}")
    print(f"Bridge workbook:    {args.bridge_xlsx.resolve()}")
    print(f"ERF pipeline JSONL: {args.erf_jsonl.resolve()}")
    print(f"Output directory:   {args.output_dir.resolve()}")
    print("Firestore operations: NONE")
    print("PSD writes:           NONE")
    print("")

    end_rows, meter_counts = read_end_rows(args.end_xlsx)
    if len(end_rows) != args.expected_end_count:
        raise SystemExit(
            f"Unexpected END row count: {len(end_rows):,}; "
            f"expected {args.expected_end_count:,}"
        )

    duplicate_meters = {
        meter: count for meter, count in meter_counts.items() if count > 1
    }
    if duplicate_meters:
        raise SystemExit(
            f"Duplicate MeterNumber groups found in END: "
            f"{len(duplicate_meters):,}"
        )

    (
        bridge_all_accounts,
        account_to_sg,
        relationship_rows,
        bridge_row_count,
    ) = read_bridge(args.bridge_xlsx)

    bridge_sg_keys = {
        sg_key
        for sg_keys in account_to_sg.values()
        for sg_key in sg_keys
    }

    if args.erf_key_path:
        erf_key_path = args.erf_key_path
        detection_candidates = []
    else:
        erf_key_path, detection_candidates = detect_erf_key_path(
            args.erf_jsonl,
            bridge_sg_keys,
        )

    print("")
    print(f"Selected ERF SG/parcel-key path: {erf_key_path}")
    if detection_candidates:
        print("Top auto-detection candidates:")
        for candidate in detection_candidates[:5]:
            print(
                "  "
                f"{candidate['path']}: "
                f"overlap={candidate['overlapCount']:,}, "
                f"nonblank={candidate['nonblank']:,}, "
                f"samples={candidate['samples']}"
            )
    print("")

    erf_key_counts, erf_key_samples, erf_record_count = read_erf_keys(
        args.erf_jsonl,
        erf_key_path,
    )

    status_counts: Counter = Counter()
    assessment_rows: list[dict[str, Any]] = []

    for index, end_row in enumerate(end_rows, start=1):
        meter_number = end_row["meterNumber"]
        account_number = end_row["accountNumber"]
        sg_keys = sorted(account_to_sg.get(account_number, set()))

        resolved_sg = ""
        pipeline_comparison_key = ""
        normalization_rule = "NONE"
        pipeline_match_count = 0

        if not account_number:
            status = "BLANK_ACCOUNT"
        elif account_number not in bridge_all_accounts:
            status = "ACCOUNT_NOT_IN_BRIDGE"
        elif not sg_keys:
            status = "ACCOUNT_HAS_NO_GIS_KEY"
        elif len(sg_keys) > 1:
            status = "ACCOUNT_MULTIPLE_SG_CODES"
        else:
            resolved_sg = sg_keys[0]
            (
                pipeline_comparison_key,
                pipeline_match_count,
                normalization_rule,
            ) = resolve_bridge_sg_to_erf(
                resolved_sg,
                erf_key_counts,
            )

            if pipeline_match_count == 0:
                status = "SG_CODE_NOT_IN_ERF_PIPELINE"
            elif pipeline_match_count > 1:
                status = "SG_CODE_MULTIPLE_ERFS"
            else:
                status = "EXACT_ONE_TO_ONE"

        status_counts[status] += 1
        assessment_rows.append(
            {
                "sourceEndRow": end_row["sourceRow"],
                "meterNumber": meter_number,
                "accountNumber": account_number,
                "distinctBridgeSgCount": len(sg_keys),
                "resolvedSgCode": resolved_sg,
                "pipelineComparisonKey": pipeline_comparison_key,
                "normalizationRule": normalization_rule,
                "pipelineMatchCount": pipeline_match_count,
                "status": status,
                "pipelineSamples": "|".join(
                    erf_key_samples.get(pipeline_comparison_key, [])
                ),
            }
        )

        if index % PROGRESS_EVERY == 0:
            print(f"[PROGRESS] Assessed {index:,} meters")

    exact_count = status_counts["EXACT_ONE_TO_ONE"]
    passed = exact_count == len(end_rows)

    assessment_csv = args.output_dir / "one_to_one_meter_assessment.csv"
    exception_csv = args.output_dir / "one_to_one_exceptions.csv"
    summary_json = args.output_dir / "one_to_one_summary.json"

    fieldnames = [
        "sourceEndRow",
        "meterNumber",
        "accountNumber",
        "distinctBridgeSgCount",
        "resolvedSgCode",
        "pipelineComparisonKey",
        "normalizationRule",
        "pipelineMatchCount",
        "status",
        "pipelineSamples",
    ]

    with assessment_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(assessment_rows)

    with exception_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(
            row
            for row in assessment_rows
            if row["status"] != "EXACT_ONE_TO_ONE"
        )

    summary = {
        "status": "PASSED" if passed else "REVIEW_REQUIRED",
        "firestoreOperations": "NONE",
        "psdWrites": "NONE",
        "sources": {
            "endXlsx": {
                "path": str(args.end_xlsx.resolve()),
                "sha256": sha256_file(args.end_xlsx),
                "rows": len(end_rows),
                "uniqueMeters": len(meter_counts),
            },
            "bridgeXlsx": {
                "path": str(args.bridge_xlsx.resolve()),
                "sha256": sha256_file(args.bridge_xlsx),
                "rows": bridge_row_count,
                "accountsPresent": len(bridge_all_accounts),
                "accountsWithGisKey": len(account_to_sg),
                "distinctGisKeys": len(bridge_sg_keys),
                "accountsWithMultipleDistinctGisKeys": sum(
                    1 for sg_keys in account_to_sg.values() if len(sg_keys) > 1
                ),
            },
            "erfJsonl": {
                "path": str(args.erf_jsonl.resolve()),
                "sha256": sha256_file(args.erf_jsonl),
                "records": erf_record_count,
                "selectedKeyPath": erf_key_path,
                "distinctNormalizedKeys": len(erf_key_counts),
                "duplicateNormalizedKeyGroups": sum(
                    1 for count in erf_key_counts.values() if count > 1
                ),
                "autoDetectionCandidates": detection_candidates,
            },
        },
        "comparisonNormalization": {
            "erfPrefixRemovedForComparison": "K241",
            "bridgeFallbackRule": "APPEND_ONE_TRAILING_ZERO",
            "sourceValuesModified": False,
        },
        "meterAssessment": {
            "totalMeters": len(end_rows),
            "statusCounts": dict(sorted(status_counts.items())),
            "normalizationRuleCounts": dict(
                sorted(
                    Counter(
                        row["normalizationRule"]
                        for row in assessment_rows
                    ).items()
                )
            ),
            "exactOneToOne": exact_count,
            "exceptions": len(end_rows) - exact_count,
        },
        "outputs": {
            "assessmentCsv": str(assessment_csv.resolve()),
            "exceptionCsv": str(exception_csv.resolve()),
        },
    }

    summary_json.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    print("")
    print("============================================================")
    print("READ-ONLY ASSESSMENT COMPLETE")
    print("============================================================")
    print(f"END meters:                  {len(end_rows):,}")
    print(f"Unique meter numbers:        {len(meter_counts):,}")
    print(f"Bridge accounts present:     {len(bridge_all_accounts):,}")
    print(f"Bridge accounts with SG:     {len(account_to_sg):,}")
    print(f"Bridge multi-SG accounts:    {sum(1 for value in account_to_sg.values() if len(value) > 1):,}")
    print(f"ERF pipeline records:        {erf_record_count:,}")
    print(f"Exact one-to-one meters:     {exact_count:,}")
    print(f"Exceptions:                  {len(end_rows) - exact_count:,}")
    print("")
    for status, count in sorted(status_counts.items()):
        print(f"{status}: {count:,}")
    print("")
    normalization_counts = Counter(
        row["normalizationRule"]
        for row in assessment_rows
    )
    print("Comparison normalization:")
    for rule, count in sorted(normalization_counts.items()):
        print(f"  {rule}: {count:,}")
    print("")
    print(f"Assessment CSV: {assessment_csv}")
    print(f"Exceptions CSV: {exception_csv}")
    print(f"Summary JSON:   {summary_json}")
    print("Firestore operations: NONE")
    print("PSD writes:           NONE")
    print(f"ASSESSMENT STATUS:    {'PASSED' if passed else 'REVIEW_REQUIRED'}")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\nAssessment interrupted.", file=sys.stderr)
        raise SystemExit(130)
