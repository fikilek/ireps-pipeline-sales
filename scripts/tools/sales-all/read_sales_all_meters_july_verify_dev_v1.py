#!/usr/bin/env python3
from __future__ import annotations
import argparse, hashlib, json, random, secrets, sys, time
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from google.cloud import firestore
from google.cloud.firestore_v1.base_query import FieldFilter
from google.oauth2 import service_account
from openpyxl import load_workbook

PROJECT_ID="ireps2"; COLLECTION="sales-all-meters"; LM="ZA5241"
JUNE="2026-06"; JULY="2026-07"
EXPECTED={"june":10216,"july":10241,"continuing":10185,"entered":56,"exited":31,"known":10272}
SOURCE_REL="input/endumeni_demo_sales/source_originals/classification/2026-07/Prepaid_Analysis_categories_30Months_Updated_July2026_v2.xlsx"
SOURCE_SHA="BB57689C32EA6F0AB1F6CE6238F2F9D61AFEB8E7590AAA62EFBD9ABA62D5A9B9"
SHEET="Prepaid_30Month_Analysis"
CREATE_IDS=("04297839708","04298618952")
EXPECTED_CATS={"CAT1":30,"CAT2":1031,"CAT3":3,"CAT4":2249,"CAT5":670,"CAT6":293,"CAT7":0,"CAT8":1320,"NORMAL":4645}
DEFAULT_SA=r"C:\dev\secrets\ireps2-e72fd9dc94de.json"

def sha(p):
    h=hashlib.sha256()
    with p.open("rb") as f:
        for b in iter(lambda:f.read(1024*1024),b""): h.update(b)
    return h.hexdigest().upper()

def entry(d,m):
    x=d.get("monthlyCategories")
    return x.get(m) if isinstance(x,dict) and isinstance(x.get(m),dict) else None

def bucket(v):
    s=str(v or "").strip().upper()
    for n in range(1,9):
        if s.startswith(f"CAT{n}"): return f"CAT{n}"
    if s.startswith("NORMAL"): return "NORMAL"
    return "MISSING"

def safe(v):
    if isinstance(v,dict): return {str(k):safe(x) for k,x in v.items()}
    if isinstance(v,list): return [safe(x) for x in v]
    if isinstance(v,datetime): return {"__type__":"timestamp","iso":v.astimezone(timezone.utc).isoformat().replace("+00:00","Z")}
    if isinstance(v,(str,int,float,bool)) or v is None: return v
    return str(v)

class P:
    def __init__(self): self.t=time.monotonic()
    def s(self,i,n,m): print(f"\n[{i}/{n}] {round(i/n*100)}% | {m} | elapsed {int(time.monotonic()-self.t)}s",flush=True)
    def h(self,m): print(f"    ... {m} | elapsed {int(time.monotonic()-self.t)}s",flush=True)

def load_source(p):
    wb=load_workbook(p,read_only=True,data_only=True); ws=wb[SHEET]
    hdr=next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
    idx={str(v).strip():i for i,v in enumerate(hdr) if v is not None}
    req=["CorrectedMeterNumber","July_2026_Category","Risk_Tier","Risk_Score","PreviousMeterNumber"]
    miss=[x for x in req if x not in idx]
    if miss: raise SystemExit(f"Missing headers: {miss}")
    out={}
    for rno,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
        mid=str(row[idx["CorrectedMeterNumber"]] or "").strip()
        if not mid: continue
        rs=row[idx["Risk_Score"]]
        rs=int(float(rs))
        if mid in out: raise SystemExit(f"Duplicate corrected meter: {mid}")
        out[mid]={"category":str(row[idx["July_2026_Category"]] or "").strip(),
                  "riskTier":str(row[idx["Risk_Tier"]] or "").strip(),
                  "riskScore":rs,
                  "previousMeterNumber":str(row[idx["PreviousMeterNumber"]] or "").strip(),
                  "sourceRow":rno}
    wb.close(); return out

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--confirm-project",required=True)
    ap.add_argument("--service-account",default=DEFAULT_SA)
    ap.add_argument("--seed",type=int)
    a=ap.parse_args()
    if a.confirm_project!=PROJECT_ID: raise SystemExit("Wrong project confirmation")
    prog=P(); prog.s(1,6,"Validate DEV target and July v2 source")
    repo=Path(__file__).resolve().parents[3]
    src=repo/SOURCE_REL
    if sha(src)!=SOURCE_SHA: raise SystemExit("July v2 source SHA mismatch")
    creds=service_account.Credentials.from_service_account_file(a.service_account)
    if getattr(creds,"project_id",None) not in (None,PROJECT_ID): raise SystemExit("Service account project mismatch")
    db=firestore.Client(project=PROJECT_ID,credentials=creds)
    if db.project!=PROJECT_ID: raise SystemExit("Firestore client project mismatch")
    print(f"    project={db.project} collection={COLLECTION} lmPcode={LM} mode=READ ONLY",flush=True)

    prog.s(2,6,"Load authoritative July v2 identity/category map")
    sm=load_source(src)
    if len(sm)!=EXPECTED["july"]: raise SystemExit(f"Source count {len(sm)} != {EXPECTED['july']}")
    prog.h(f"source identities {len(sm):,}")

    prog.s(3,6,"Scan actual ZA5241 Firestore month state")
    docs={}; cats=Counter(); scanned=0
    q=db.collection(COLLECTION).where(filter=FieldFilter("lmPcode","==",LM)).select(["monthlyCategories","metadata","previousMeterNumber"])
    for snap in q.stream():
        scanned+=1; d=snap.to_dict() or {}; docs[snap.id]=d
        je=entry(d,JULY)
        if je: cats[bucket(je.get("leakageCategory"))]+=1
        if scanned%500==0: prog.h(f"scanned {scanned:,}")
    june={i for i,d in docs.items() if entry(d,JUNE)}
    july={i for i,d in docs.items() if entry(d,JULY)}
    cont=june&july; entered=july-june; exited=june-july; known=june|july
    prog.h(f"June={len(june):,} July={len(july):,} continuing={len(cont):,} entered={len(entered)} exited={len(exited)}")

    prog.s(4,6,"Compare all July Firestore categories to source")
    missing=[]; mism=[]; meta=[]
    reqmeta={"createdAt","createdByUid","createdByUser","updatedAt","updatedByUid","updatedByUser"}
    for n,(mid,s) in enumerate(sorted(sm.items()),1):
        d=docs.get(mid)
        if d is None: missing.append(mid); continue
        actual=entry(d,JULY)
        exp={"leakageCategory":s["category"],"riskTier":s["riskTier"],"riskScore":s["riskScore"]}
        got={k:actual.get(k) for k in exp} if actual else None
        if got!=exp: mism.append({"meterId":mid,"expected":exp,"actual":safe(got)})
        md=d.get("metadata")
        if not isinstance(md,dict) or set(md)!=reqmeta: meta.append(mid)
        if n%1000==0: prog.h(f"compared {n:,}/{len(sm):,}")

    prog.s(5,6,"Build random samples, creates and replacement pairs")
    rng=random.Random(a.seed if a.seed is not None else secrets.randbits(64)); seed=rng.getrandbits(64)
    reasons=[]; byb=defaultdict(list)
    for mid in sorted(july):
        byb[bucket(entry(docs[mid],JULY).get("leakageCategory"))].append(mid)
    for b in [*(f"CAT{n}" for n in range(1,9))]:
        ids=byb[b];
        for mid in rng.sample(ids,min(3,len(ids))): reasons.append((f"{b} random",mid))
    for mid in rng.sample(byb["NORMAL"],min(5,len(byb["NORMAL"]))): reasons.append(("NORMAL random",mid))
    for mid in CREATE_IDS: reasons.append(("JULY CREATE",mid))
    for mid in rng.sample(sorted(cont),3): reasons.append(("CONTINUING June+July",mid))
    entered_existing=sorted(set(entered)-set(CREATE_IDS))
    for mid in rng.sample(entered_existing,3): reasons.append(("ENTERED July existing",mid))
    pairs=[]
    cands=[]
    for suc in sorted(entered):
        prev=str(docs.get(suc,{}).get("previousMeterNumber") or "").strip()
        if prev in exited: cands.append((suc,prev))
    for suc,prev in rng.sample(cands,min(2,len(cands))):
        pairs.append({"successor":suc,"predecessor":prev,"successorHasJune":suc in june,"successorHasJuly":suc in july,"predecessorHasJune":prev in june,"predecessorHasJuly":prev in july})
        reasons += [("REPLACEMENT successor",suc),("REPLACEMENT predecessor",prev)]
    dedup={}
    for reason,mid in reasons: dedup.setdefault(mid,reason)
    samples=[]
    for n,(mid,reason) in enumerate(dedup.items(),1):
        s=db.collection(COLLECTION).document(mid).get(); d=s.to_dict() or {}
        samples.append({"sampleReason":reason,"documentId":mid,
                        "createTime":s.create_time.astimezone(timezone.utc).isoformat().replace("+00:00","Z") if s.create_time else None,
                        "updateTime":s.update_time.astimezone(timezone.utc).isoformat().replace("+00:00","Z") if s.update_time else None,
                        "document":safe(d)})
        prog.h(f"sample {n}/{len(dedup)} {reason} {mid}")

    creates={}
    for mid in CREATE_IDS:
        s=db.collection(COLLECTION).document(mid).get(); d=s.to_dict() or {}
        creates[mid]={"exists":s.exists,"hasJuneCategory":bool(entry(d,JUNE)),"hasJulyCategory":bool(entry(d,JULY)),
                      "monthlySalesC":safe(d.get("monthlySalesC")),"monthlyTotalsC":safe(d.get("monthlyTotalsC")),"monthlyUnits":safe(d.get("monthlyUnits"))}

    prog.s(6,6,"Write independent evidence")
    checks={
      "sourceHash":sha(src)==SOURCE_SHA,"sourceCount":len(sm)==EXPECTED["july"],
      "june":len(june)==EXPECTED["june"],"july":len(july)==EXPECTED["july"],
      "continuing":len(cont)==EXPECTED["continuing"],"entered":len(entered)==EXPECTED["entered"],
      "exited":len(exited)==EXPECTED["exited"],"known":len(known)==EXPECTED["known"],
      "missing":not missing,"categoryMismatches":not mism,"metadataFailures":not meta,
      "categoryCounts":all(cats.get(k,0)==v for k,v in EXPECTED_CATS.items()),
      "create1":creates[CREATE_IDS[0]]["exists"] and creates[CREATE_IDS[0]]["hasJulyCategory"],
      "create2":creates[CREATE_IDS[1]]["exists"] and creates[CREATE_IDS[1]]["hasJulyCategory"],
    }
    verdict="INDEPENDENT STAGE 9 STEP 4 JULY DEV STATE VERIFIED" if all(checks.values()) else "INDEPENDENT STAGE 9 STEP 4 JULY DEV STATE FAILED"
    result={"verdict":verdict,"readOnly":True,"projectId":PROJECT_ID,"collection":COLLECTION,"lmPcode":LM,
            "julySourceSha256":sha(src),"scopeScanned":scanned,"sourceIdentityCount":len(sm),
            "juneActiveCount":len(june),"julyActiveCount":len(july),"continuingCount":len(cont),
            "enteredCount":len(entered),"exitedCount":len(exited),"knownCount":len(known),
            "categoryCounts":dict(cats),"missingJulySourceIds":missing,"julyCategoryMismatches":mism,
            "metadataFailures":meta,"enteredIds":sorted(entered),"exitedIds":sorted(exited),
            "replacementPairs":pairs,"createChecks":creates,"randomSeed":seed,"sampleRecords":samples,
            "checks":checks,"firestoreWritesPerformed":0}
    out=repo/"output"/"logs"; stamp=datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    jp=out/f"independent_stage9_step4_july_dev__{stamp}.json"
    mp=out/f"independent_stage9_step4_july_dev__{stamp}.md"
    jp.write_text(json.dumps(result,indent=2,ensure_ascii=False),encoding="utf-8")
    lines=[f"# {verdict}","",f"- Project: `{PROJECT_ID}`",f"- Collection: `{COLLECTION}`",f"- Scope scanned: **{scanned:,}**",
           f"- June Active: **{len(june):,}**",f"- July Active: **{len(july):,}**",f"- Continuing: **{len(cont):,}**",
           f"- Entered: **{len(entered)}**",f"- Exited: **{len(exited)}**",f"- Known: **{len(known):,}**",
           f"- Source category mismatches: **{len(mism)}**",f"- Metadata failures: **{len(meta)}**","",
           "## Category counts","",json.dumps(dict(cats),indent=2),"","## Two creates","",json.dumps(creates,indent=2),
           "","## Replacement pairs","",json.dumps(pairs,indent=2),"","## Sampled full documents",""]
    for r in samples:
        lines += [f"### {r['sampleReason']} — `{r['documentId']}`","```json",json.dumps(r,indent=2,ensure_ascii=False),"```",""]
    mp.write_text("\n".join(lines),encoding="utf-8")
    print("\n============================================================")
    print(verdict)
    print("============================================================")
    print(f"Scope scanned : {scanned:,}")
    print(f"June Active   : {len(june):,}")
    print(f"July Active   : {len(july):,}")
    print(f"Continuing    : {len(cont):,}")
    print(f"Entered       : {len(entered)}")
    print(f"Exited        : {len(exited)}")
    print(f"Known         : {len(known):,}")
    print(f"Category mismatches: {len(mism)}")
    print(f"Metadata failures : {len(meta)}")
    print(f"JSON          : {jp}")
    print(f"Markdown      : {mp}")
    print("Firestore writes performed: 0")
    if not all(checks.values()):
        print("FAILED CHECKS:", [k for k,v in checks.items() if not v])
        return 1
    return 0

if __name__=="__main__":
    raise SystemExit(main())
